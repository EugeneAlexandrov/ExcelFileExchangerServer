# Файл views.py
import psycopg2
import openpyxl
from django.core.files.storage import FileSystemStorage
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.backends.db import SessionStore
import datetime as dt
from io import BytesIO
from django.http import FileResponse

from .database_requests import (getDataTableForAllTime,
                                getUserInfoFromDB,
                                getUserInfoFromDBDataset,
                                getTranzitUserInfoFromDB,
                                getAllTranzitUserInfoFromDB,
                                getMaxWarehouseQty,
                                getMaxWarehouseAllQty,
                                getNormsWarehouseQty,
                                getNormsWarehouseAllQty)
from .models import DailyMonitoringUserData,ConstantUserData

DAYDELTA = dt.timedelta(days=1,
                           seconds=0,
                           microseconds=0,
                           milliseconds=0,
                           minutes=0,
                           hours=0,
                           weeks=0)


@login_required(login_url='')
def home(request):
    #В данном случае ID админа 2, но рекомендуется использовать 1, и первым пользователем создавать именно админа.
    if request.user.id != 1:
        if request.method == 'POST':
            file = request.FILES.get('file')
            if file is not None:
                if file.name.endswith('.xlsx'):
                    fs = FileSystemStorage()
                    filename = fs.save(file.name, file)
                    try:
                        # Чтение xlsx файла
                        df = pd.read_excel(fs.path(filename))
                        # Игнорирование первых трёх строк
                        df = df.iloc[2:]
                        # Преобразование DataFrame в списки столбцов (В дальнейшем для новых столбцов таблицы создавать новые поля в таблице БД, а также добавлять ниже в переменные)
                        ImportIn = df.iloc[:, 0].tolist()                     # в т.ч. импорт Прибыло
                        ImportOut = df.iloc[:, 1].tolist()                    # в т.ч. импорт Убыло
                        ExportIn = df.iloc[:, 2].tolist()                     # в т.ч. экспорт Прибыло
                        ExportOut = df.iloc[:, 3].tolist()                    # в т.ч. экспорт Убыло
                        TransitIn = df.iloc[:, 4].tolist()                    # в т.ч. транзит Прибыло
                        TransitOut = df.iloc[:, 5].tolist()                   # в т.ч. транзит Убыло
                        ExportEmpty = df.iloc[:, 6].tolist()                  # в т.ч. экспорт порожние
                        OtherEmpty = df.iloc[:, 7].tolist()                   # в т.ч. прочие порожние
                        UnloadReid = df.iloc[:, 8].tolist()                   # На рейде в ожидании Выгрузки
                        LoadingReid = df.iloc[:, 9].tolist()                  # На рейде в ожидании Погрузки
                        UnloadPort = df.iloc[:, 10].tolist()                  # На подходах к порту для Выгрузки
                        LoadingPort = df.iloc[:, 11].tolist()                 # На подходах к порту для Погрузки
                        fs.delete(filename)

                        # Дальнейшая обработка данных
                        request.session['parameters'] = {
                                                            'ImportIn': ImportIn,
                                                            'ImportOut': ImportOut,
                                                            'ExportIn': ExportIn,
                                                            'ExportOut': ExportOut,
                                                            'TransitIn': TransitIn,
                                                            'TransitOut': TransitOut,
                                                            'ExportEmpty': ExportEmpty,
                                                            'OtherEmpty': OtherEmpty,
                                                            'UnloadReid': UnloadReid,
                                                            'LoadingReid': LoadingReid,
                                                            'UnloadPort': UnloadPort,
                                                            'LoadingPort': LoadingPort
                                                            }
                        # Сохраните сессию, чтобы сгенерировать сессионный ключ
                        request.session.save()

                        # Получить текущий session ID
                        session_id = request.session.session_key
                        # Создайте URL-адрес перенаправления с этим session ID
                        redirect_url = f'/confirm/?session_id={session_id}'
                        return redirect(redirect_url)
                    except:
                        fs.delete(filename)
                        return render(request, 'error.html', {'ErrorText' : 'Ошибка выгрузки данных'})
                else:
                    return render(request, 'error.html', {'ErrorText' : 'Неверный формат файла'})
            else:
                return redirect('confirm')
        return render(request, 'home.html')
    else:
        return render(request, 'admin.html')

@login_required(login_url='')
def register(request):
    if request.user.id == 1:
        if request.method == 'POST':
            try:
                username = request.POST['username']
                password = request.POST['password']
                norms = request.POST['norms']
                max = request.POST['max']

                # Создание нового пользователя
                user = User.objects.create_user(username=username, password=password)
                user.save()
                ConstantUserData.objects.create(db_userid=user.id, db_norms=norms, db_max=max)
                return redirect('login')
            except:
                return render(request, 'register.html')
        return render(request, 'register.html')
    else:
        return redirect('login')
    ####

def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        # Проверка введенных данных
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            return redirect('login')
    return render(request, 'login.html')

def user_logout(request):
    # Выход пользователя
    logout(request)
    return redirect('home')

@login_required(login_url='')
def confirm(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        ImportIn = params.get('ImportIn')
        ImportOut = params.get('ImportOut')
        ExportIn = params.get('ExportIn')
        ExportOut = params.get('ExportOut')
        TransitIn = params.get('TransitIn')
        TransitOut = params.get('TransitOut')
        ExportEmpty = params.get('ExportEmpty')
        OtherEmpty = params.get('OtherEmpty')
        UnloadReidlin = params.get('UnloadReidlin')
        UnloadReidtramp = params.get('UnloadReidtramp')
        LoadingReidLin = params.get('LoadingReidLin')
        LoadingReidTramp = params.get('LoadingReidTramp')
        UnloadPortlin = params.get('UnloadPortlin')
        UnloadPorttramp = params.get('UnloadPorttramp')
        LoadingPortLin = params.get('LoadingPortLin')
        LoadingPortTramp = params.get('LoadingPortTramp')
        if request.method == 'POST':
            date2 = [[request.POST['date2']]]
            ImportIn = [request.POST['ImportIn']]
            ImportOut = [request.POST['ImportOut']]
            ExportIn = [request.POST['ExportIn']]
            ExportOut = [request.POST['ExportOut']]
            TransitIn = [request.POST['TransitIn']]
            TransitOut = [request.POST['TransitOut']]
            ExportEmpty = [request.POST['ExportEmpty']]
            OtherEmpty = [request.POST['OtherEmpty']]
            UnloadReidlin = [request.POST['UnloadReidlin']]
            UnloadReidtramp = [request.POST['UnloadReidtramp']]
            LoadingReidLin = [request.POST['LoadingReidLin']]
            LoadingReidTramp = [request.POST['LoadingReidTramp']]
            UnloadPortlin = [request.POST['UnloadPortlin']]
            UnloadPorttramp = [request.POST['UnloadPorttramp']]
            LoadingPortLin = [request.POST['LoadingPortLin']]
            LoadingPortTramp = [request.POST['LoadingPortTramp']]
            request.session['parameters'] = {
                'date2' : date2,
                'ImportIn': ImportIn,
                'ImportOut': ImportOut,
                'ExportIn': ExportIn,
                'ExportOut': ExportOut,
                'TransitIn': TransitIn,
                'TransitOut': TransitOut,
                'ExportEmpty': ExportEmpty,
                'OtherEmpty': OtherEmpty,
                'UnloadReidlin': UnloadReidlin,
                'UnloadReidtramp': UnloadReidtramp,
                'LoadingReidLin': LoadingReidLin,
                'LoadingReidTramp': LoadingReidTramp,
                'UnloadPortlin' : UnloadPortlin,
                'UnloadPorttramp': UnloadPorttramp,
                'LoadingPortLin': LoadingPortLin,
                'LoadingPortTramp': LoadingPortTramp
            }
            redirect_url = f'/success/?session_id={session_id}'
            return redirect(redirect_url)
        return render(request, 'confirm.html', {
                                                        'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                                                        'ImportIn': ImportIn[0],
                                                        'ImportOut': ImportOut[0],
                                                        'ExportIn': ExportIn[0],
                                                        'ExportOut': ExportOut[0],
                                                        'TransitIn': TransitIn[0],
                                                        'TransitOut': TransitOut[0],
                                                        'ExportEmpty': ExportEmpty[0],
                                                        'OtherEmpty': OtherEmpty[0],
                                                        'UnloadReidlin': UnloadReidlin[0],
                                                        'UnloadReidtramp': UnloadReidtramp[0],
                                                        'LoadingReidLin': LoadingReidLin[0],
                                                        'LoadingReidTramp': LoadingReidTramp[0],
                                                        'UnloadPortlin' : UnloadPortlin[0],
                                                        'UnloadPorttramp': UnloadPorttramp[0],
                                                        'LoadingPortLin': LoadingPortLin[0],
                                                        'LoadingPortTramp': LoadingPortTramp[0]
        })
    else:
        if request.method == 'POST':
            date2 = [request.POST['date2']],
            ImportIn = [request.POST['ImportIn']]
            ImportOut = [request.POST['ImportOut']]
            ExportIn = [request.POST['ExportIn']]
            ExportOut = [request.POST['ExportOut']]
            TransitIn = [request.POST['TransitIn']]
            TransitOut = [request.POST['TransitOut']]
            ExportEmpty = [request.POST['ExportEmpty']]
            OtherEmpty = [request.POST['OtherEmpty']]
            UnloadReidlin = [request.POST['UnloadReidlin']]
            UnloadReidtramp = [request.POST['UnloadReidtramp']]
            LoadingReidLin = [request.POST['LoadingReidLin']]
            LoadingReidTramp = [request.POST['LoadingReidTramp']]
            UnloadPortlin = [request.POST['UnloadPortlin']]
            UnloadPorttramp = [request.POST['UnloadPorttramp']]
            LoadingPortLin = [request.POST['LoadingPortLin']]
            LoadingPortTramp = [request.POST['LoadingPortTramp']]
            request.session['parameters'] = {
                'date2': date2,
                'ImportIn': ImportIn,
                'ImportOut': ImportOut,
                'ExportIn': ExportIn,
                'ExportOut': ExportOut,
                'TransitIn': TransitIn,
                'TransitOut': TransitOut,
                'ExportEmpty': ExportEmpty,
                'OtherEmpty' : OtherEmpty,
                'UnloadReidlin': UnloadReidlin,
                'UnloadReidtramp': UnloadReidtramp,
                'LoadingReidLin': LoadingReidLin,
                'LoadingReidTramp': LoadingReidTramp,
                'UnloadPortlin' : UnloadPortlin,
                'UnloadPorttramp': UnloadPorttramp,
                'LoadingPortLin': LoadingPortLin,
                'LoadingPortTramp': LoadingPortTramp
            }
            # Сохраните сессию, чтобы сгенерировать сессионный ключ
            request.session.save()

            # Получить текущий session ID
            session_id = request.session.session_key
            # Создайте URL-адрес перенаправления с этим session ID
            redirect_url = f'/success/?session_id={session_id}'
            return redirect(redirect_url)
        try:
            #WORK IN PROGRESS
            Userdata = getUserInfoFromDB(request.user.id, (dt.datetime.now()).strftime('%Y-%m-%d'))
            return render(request, 'confirm.html', {
                'date2' : (dt.datetime.now()).strftime('%Y-%m-%d'),
                'ImportIn': Userdata[0][0],
                'ImportOut': Userdata[0][1],
                'ExportIn': Userdata[0][2],
                'ExportOut': Userdata[0][3],
                'TransitIn': Userdata[0][4],
                'TransitOut': Userdata[0][5],
                'ExportEmpty': Userdata[0][6],
                'OtherEmpty': Userdata[0][7],
                'UnloadReidlin': Userdata[0][8],
                'UnloadReidtramp': Userdata[0][9],
                'LoadingReidLin': Userdata[0][10],
                'LoadingReidTramp': Userdata[0][11],
                'UnloadPortlin': Userdata[0][12],
                'UnloadPorttramp': Userdata[0][13],
                'LoadingPortLin': Userdata[0][13],
                'LoadingPortTramp': Userdata[0][14]
            })
        except:
            return render(request, 'confirm.html', {
                'date2': (dt.datetime.now()).strftime('%Y-%m-%d'),
                'ImportIn': 0,
                'ImportOut': 0,
                'ExportIn': 0,
                'ExportOut': 0,
                'TransitIn': 0,
                'TransitOut': 0,
                'ExportEmpty': 0,
                'OtherEmpty': 0,
                'UnloadReidlin': 0,
                'UnloadReidtramp': 0,
                'LoadingReidLin': 0,
                'LoadingReidTramp': 0,
                'UnloadPortlin': 0,
                'UnloadPorttramp': 0,
                'LoadingPortLin': 0,
                'LoadingPortTramp': 0
            })

@login_required(login_url='')
def success(request):
    session_id = request.GET.get('session_id')
    if session_id:
        # Используйте session_id, чтобы вручную загрузить сеанс
        request.session = SessionStore(session_key=session_id)
        params = request.session.get('parameters',{})
        date2 = params.get('date2')
        print(date2)
        ImportIn = params.get('ImportIn')
        ImportOut = params.get('ImportOut')
        ExportIn = params.get('ExportIn')
        ExportOut = params.get('ExportOut')
        TransitIn = params.get('TransitIn')
        TransitOut = params.get('TransitOut')
        ExportEmpty = params.get('ExportEmpty')
        OtherEmpty = params.get('OtherEmpty')
        UnloadReidlin = params.get('UnloadReidlin')
        UnloadReidtramp = params.get('UnloadReidtramp')
        LoadingReidLin = params.get('LoadingReidLin')
        LoadingReidTramp = params.get('LoadingReidTramp')
        UnloadPortlin = params.get('UnloadPortlin')
        UnloadPorttramp = params.get('UnloadPorttramp')
        LoadingPortLin = params.get('LoadingPortLin')
        LoadingPortTramp = params.get('LoadingPortTramp')
        ImportIn[0] = NanCheck(ImportIn[0])
        ImportOut[0] = NanCheck(ImportOut[0])
        ExportIn[0] = NanCheck(ExportIn[0])
        ExportOut[0] = NanCheck(ExportOut[0])
        TransitIn[0] = NanCheck(TransitIn[0])
        TransitOut[0] = NanCheck(TransitOut[0])
        ExportEmpty[0] = NanCheck(ExportEmpty[0])
        OtherEmpty[0] = NanCheck(OtherEmpty[0])
        UnloadReidlin[0] = NanCheck(UnloadReidlin[0])
        UnloadReidtramp[0] = NanCheck(UnloadReidtramp[0])
        LoadingReidLin[0] = NanCheck(LoadingReidLin[0])
        LoadingReidTramp[0] = NanCheck(LoadingReidTramp[0])
        UnloadPortlin[0] = NanCheck(UnloadPortlin[0])
        UnloadPorttramp[0] = NanCheck(UnloadPorttramp[0])
        LoadingPortLin[0] = NanCheck(LoadingPortLin[0])
        LoadingPortTramp[0] = NanCheck(LoadingPortTramp[0])
        # Перезапись данных за предыдущий день, при совпадении даты и ID пользователя.
        DataItem = DailyMonitoringUserData.objects.filter(date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'), db_userid = request.user.id).update(
                db_importin=int(ImportIn[0]),
                db_importout=int(ImportOut[0]),
                db_exportin=int(ExportIn[0]),
                db_exportout=int(ExportOut[0]),
                db_transitin=int(TransitIn[0]),
                db_transitout=int(TransitOut[0]),
                db_exportempty=int(ExportEmpty[0]),
                db_otherempty=int(OtherEmpty[0]),
                db_unload_reid_lin=int(UnloadReidlin[0]),
                db_unload_reid_tramp=int(UnloadReidtramp[0]),
                db_loading_reid_lin=int(LoadingReidLin[0]),
                db_loading_reid_tramp=int(LoadingReidTramp[0]),
                db_loading_port_lin=int(UnloadPortlin[0]),
                db_loading_port_tramp=int(UnloadPorttramp[0]),
                db_unload_port_lin=int(LoadingPortLin[0]),
                db_unload_port_tramp=int(LoadingPortTramp[0])
            )
        # Запись новых данных, если ID пользователя и дата не совпадают.
        if DataItem == 0:
            DailyMonitoringUserData.objects.create(date = dt.datetime.strptime(date2[0][0], '%Y-%m-%d'),
                                          db_userid = request.user.id,
                                          db_importin= int(ImportIn[0]),
                                          db_importout = int(ImportOut[0]),
                                          db_exportin = int(ExportIn[0]),
                                          db_exportout = int(ExportOut[0]),
                                          db_transitin = int(TransitIn[0]),
                                          db_transitout = int(TransitOut[0]),
                                          db_exportempty = int(ExportEmpty[0]),
                                          db_otherempty = int(OtherEmpty[0]),
                                          db_unload_reid_lin=int(UnloadReidlin[0]),
                                          db_unload_reid_tramp=int(UnloadReidtramp[0]),
                                          db_loading_reid_lin=int(LoadingReidLin[0]),
                                          db_loading_reid_tramp=int(LoadingReidTramp[0]),
                                          db_loading_port_lin=int(UnloadPortlin[0]),
                                          db_loading_port_tramp=int(UnloadPorttramp[0]),
                                          db_unload_port_lin=int(LoadingPortLin[0]),
                                          db_unload_port_tramp=int(LoadingPortTramp[0])
                                          )

        return render(request, 'success.html', {
                                                'date2': date2[0][0],
                                                'ImportIn': ImportIn[0],
                                                'ImportOut': ImportOut[0],
                                                'ExportIn': ExportIn[0],
                                                'ExportOut': ExportOut[0],
                                                'TransitIn': TransitIn[0],
                                                'TransitOut': TransitOut[0],
                                                'ExportEmpty': ExportEmpty[0],
                                                'OtherEmpty': OtherEmpty[0],
                                                'UnloadReidlin': UnloadReidlin[0],
                                                'UnloadReidtramp': UnloadReidtramp[0],
                                                'LoadingReidLin': LoadingReidLin[0],
                                                'LoadingReidTramp': LoadingReidTramp[0],
                                                'UnloadPortlin' : UnloadPortlin[0],
                                                'UnloadPorttramp': UnloadPorttramp[0],
                                                'LoadingPortLin': LoadingPortLin[0],
                                                'LoadingPortTramp': LoadingPortTramp[0],
                                                'user': request.user})

def NanCheck(i):
    try:
        float(i)
        if str(i) == 'nan' or str(i) =='':
            i = 0
        return i
    except:
        i = 0
    return i

# ###File download realization
# @login_required(login_url='')
# def download(request):
#     if request.user.id == 1:
#         wb = openpyxl.load_workbook('./Test.xlsx')
#         ws = wb.get_sheet_by_name('Шаблон1')
#         params = request.session.get('parameters', {})
#         date = params.get('date1')
#         try:
#             #admin_id = 2
#             #
#             # TestUser1
#             #
#             User1data = getUserInfoFromDB(2, date)
#             #ws[f'A2'] = date
#             ws[f'A4'] = 'TestUser1'
#             ws[f'B4'] = User1data[0][0]
#             ws[f'C4'] = User1data[0][1]
#             ws[f'D4'] = User1data[0][2]
#             ws[f'E4'] = User1data[0][3]
#             ws[f'F4'] = User1data[0][4]
#             ws[f'G4'] = User1data[0][5]
#             ws[f'H4'] = User1data[0][6]
#             ws[f'I4'] = User1data[0][7]
#             ws[f'J4'] = User1data[0][8]
#             ws[f'K4'] = User1data[0][9]
#             ws[f'L4'] = User1data[0][10]
#             ws[f'M4'] = User1data[0][11]
#
#             #
#             # TestUser2
#             #
#
#             User2data = getUserInfoFromDB(3, date)
#
#             #ws[f'A3'] = date
#             ws[f'A5'] = 'TestUser1'
#             ws[f'B5'] = User2data[0][0]
#             ws[f'C5'] = User2data[0][1]
#             ws[f'D5'] = User2data[0][2]
#             ws[f'E5'] = User2data[0][3]
#             ws[f'F5'] = User2data[0][4]
#             ws[f'G5'] = User2data[0][5]
#             ws[f'H5'] = User2data[0][6]
#             ws[f'I5'] = User2data[0][7]
#             ws[f'J5'] = User2data[0][8]
#             ws[f'K5'] = User2data[0][9]
#             ws[f'L5'] = User2data[0][10]
#             ws[f'M5'] = User2data[0][11]
#
#             #
#             # TestUser3 massive query
#             #
#             User3data = getUserInfoFromDB(4, date)
#
#             #ws[f'A4'] = date
#             ws[f'A6'] = 'TestUser1'
#             ws[f'B6'] = User2data[0][0]
#             ws[f'C6'] = User2data[0][1]
#             ws[f'D6'] = User2data[0][2]
#             ws[f'E6'] = User2data[0][3]
#             ws[f'F6'] = User2data[0][4]
#             ws[f'G6'] = User2data[0][5]
#             ws[f'H6'] = User2data[0][6]
#             ws[f'I6'] = User2data[0][7]
#             ws[f'J6'] = User2data[0][8]
#             ws[f'K6'] = User2data[0][9]
#             ws[f'L6'] = User2data[0][10]
#             ws[f'M6'] = User2data[0][11]
#             #
#             # Сумма по всем пользователям за дату
#             #
#             UserAlldatafordate = getDataTableForDate(date)
#             #ws[f'A5'] = date
#             ws[f'A7'] = 'TestUser1'
#             ws[f'B7'] = UserAlldatafordate[0][0]
#             ws[f'C7'] = UserAlldatafordate[0][1]
#             ws[f'D7'] = UserAlldatafordate[0][2]
#             ws[f'E7'] = UserAlldatafordate[0][3]
#             ws[f'F7'] = UserAlldatafordate[0][4]
#             ws[f'G7'] = UserAlldatafordate[0][5]
#             ws[f'H7'] = UserAlldatafordate[0][6]
#             ws[f'I7'] = UserAlldatafordate[0][7]
#             ws[f'J7'] = UserAlldatafordate[0][8]
#             ws[f'K7'] = UserAlldatafordate[0][9]
#             ws[f'L7'] = UserAlldatafordate[0][10]
#             ws[f'M7'] = UserAlldatafordate[0][11]
#             #
#             # Сумма по всем пользователям за всё время
#             #
#             UserAlldata = getDataTableForAllTime(date)
#             #ws[f'A6'] = date
#             ws[f'A8'] = 'TestUser1'
#             ws[f'B8'] = UserAlldata[0][0]
#             ws[f'C8'] = UserAlldata[0][1]
#             ws[f'D8'] = UserAlldata[0][2]
#             ws[f'E8'] = UserAlldata[0][3]
#             ws[f'F8'] = UserAlldata[0][4]
#             ws[f'G8'] = UserAlldata[0][5]
#             ws[f'H8'] = UserAlldata[0][6]
#             ws[f'I8'] = UserAlldata[0][7]
#             ws[f'J8'] = UserAlldata[0][8]
#             ws[f'K8'] = UserAlldata[0][9]
#             ws[f'L8'] = UserAlldata[0][10]
#             ws[f'M8'] = UserAlldata[0][11]
#         except:
#             pass
#         bytes_io = BytesIO()
#         wb.save(bytes_io)
#         bytes_io.seek(0)
#         return FileResponse(bytes_io, as_attachment=True, filename=(date+'.xslx'))
#     else:
#         return redirect('home')

###File download realization
@login_required(login_url='')
def download(request):
    if request.user.id == 1:
        wb = openpyxl.load_workbook('./WS1_s.xlsx')
        ws = wb.get_sheet_by_name('Шаблон1')
        params = request.session.get('parameters', {})
        date = params.get('date1')
        try:
            #admin_id = 2
            #
            # TestUser1
            #
            User1data = getUserInfoFromDB(2, date)
            #
            # TestUser2
            #
            User2data = getUserInfoFromDB(3, date)
            #
            # TestUser3 massive query
            #
            User3data = getUserInfoFromDB(4, date)
            #
            # Сумма по всем пользователям за дату
            #
            #
            # Сумма по всем пользователям за всё время
            #
            UserAllDataOlder = getDataTableForAllTime((dt.datetime.strptime(date, '%Y-%m-%d')-DAYDELTA).strftime('%Y-%m-%d'))
            UserAlldata = getDataTableForAllTime(date)

            ws['E9'] = (UserAlldata[0][0]-UserAlldata[0][1]+UserAlldata[0][2]-UserAlldata[0][3]+UserAlldata[0][4]-UserAlldata[0][5])
            ws['Q22'] = UserAlldata[0][0]
            ws['R22'] = UserAlldata[0][1]
            ws['S22'] = UserAlldata[0][2]
            ws['T22'] = UserAlldata[0][3]
            ws['U22'] = UserAlldata[0][4]
            ws['V22'] = UserAlldata[0][5]
            ws['W22'] = UserAlldata[0][6]
            ws['X22'] = UserAlldata[0][7]
            ws['Y22'] = UserAlldata[0][8]
            ws['Z22'] = UserAlldata[0][9]
            ws['AA22'] = UserAlldata[0][10]
            ws['AB22'] = UserAlldata[0][11]
            ws['C33'] = UserAlldata[0][8]
            ws['D33'] = UserAlldata[0][9]
            ws['F33'] = UserAlldata[0][10]
            ws['G33'] = UserAlldata[0][11]
        except:
            pass
        bytes_io = BytesIO()
        wb.save(bytes_io)
        bytes_io.seek(0)
        return FileResponse(bytes_io, as_attachment=True, filename=(date+'.xslx'))
    else:
        return redirect('home')

@login_required(login_url='')
def dataset(request):
    if request.user.id == 1:
        params = request.session.get('parameters', {})
        date = params.get('date1')
        # try:
            #
            # TestUser1Table1
            #
        Tranzit1 = getTranzitUserInfoFromDB(2,date)
        User1data = getUserInfoFromDBDataset(2, date)
        MaxWarehouseQty1User = getMaxWarehouseQty(2)[0][0]
        NormsWarehouseQty1User = getNormsWarehouseQty(2)[0][0]
        AllQty1User = AllQtyCalculator(User1data,Tranzit1)
        AllQtyPercent1User = AllQtyPercent(AllQty1User,MaxWarehouseQty1User,0)

        #
        # TestUser2Table1
        #

        Tranzit2 = getTranzitUserInfoFromDB(3,date)
        User2data = getUserInfoFromDBDataset(3, date)
        MaxWarehouseQty2User = getMaxWarehouseQty(3)[0][0]
        NormsWarehouseQty2User = getNormsWarehouseQty(3)[0][0]
        AllQty2User = AllQtyCalculator(User2data,Tranzit2)
        AllQtyPercent2User = AllQtyPercent(AllQty2User,MaxWarehouseQty2User,0)



        UserAlldata = getDataTableForAllTime(date)
        TranzitAll = getAllTranzitUserInfoFromDB(date)
        MaxWarehouseQtyAll = getMaxWarehouseAllQty()[0][0]
        NormsWarehouseQtyAll = getNormsWarehouseAllQty()[0][0]
        AllQtyAll = AllQtyCalculator(UserAlldata,TranzitAll)
        AllQtyPercent1All = AllQtyPercent(AllQtyAll,MaxWarehouseQtyAll,0)
        # except:
        #     return render(request, 'error.html', {'ErrorText' : 'Ошибка отображения данных'})
        return render(request, 'dataset.html', {
                                                    'date' : ((dt.datetime.strptime(date,'%Y-%m-%d')+DAYDELTA).strftime('%d.%m.%Y.')),

                                                    'User1data' : User1data[0],
                                                    'User2data': User2data[0],
                                                    'UserAlldata': UserAlldata[0],

                                                    'AllQtyPercent1User': AllQtyPercent1User,
                                                    'AllQty1User' :AllQty1User,
                                                    'MaxWarehouseQty2User': MaxWarehouseQty2User,
                                                    'NormsWarehouseQty1User' : NormsWarehouseQty1User,

                                                    'AllQtyPercent2User': AllQtyPercent2User,
                                                    'AllQty2User': AllQty2User,
                                                    'MaxWarehouseQty1User': MaxWarehouseQty1User,
                                                    'NormsWarehouseQty2User': NormsWarehouseQty2User,

                                                    'AllQtyAll' :AllQtyAll,
                                                    'AllQtyPercent1All' : AllQtyPercent1All,
                                                    'MaxWarehouseQtyAll':MaxWarehouseQtyAll,
                                                    'NormsWarehouseQtyAll':NormsWarehouseQtyAll,

                                                    'user': request.user})
    else:
        return redirect('home')

@login_required(login_url='')
def datepick_admin(request):
    if request.user.id == 1:
        if request.method == 'POST':
            date1 = request.POST['date1']
            request.session['parameters'] = {'date1': ((dt.datetime.strptime(date1, '%Y-%m-%d') - DAYDELTA).strftime('%Y-%m-%d'))}
            return redirect(dataset)
        return render(request, 'datepick_admin.html' , {'currentdate':(dt.datetime.now()-DAYDELTA).strftime('%Y-%m-%d')})
    else:
        return redirect('home')


def set_border(ws, cell_range):
    thin = openpyxl.styles.Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

def AllQtyPercent(top,bot,signs):
    # For persents with signs after , 0 if no signs
    if (top != None and bot != None):
        return int(round((top/bot*100),signs))
    else:
        return 0


def AllQtyCalculator(UserInfo,TranzitInfo):
    if (UserInfo[0][0] != None
            and UserInfo[0][1] != None
            and UserInfo[0][2] != None
            and UserInfo[0][3] != None):
                if TranzitInfo[0][0] != None :
                    return (UserInfo[0][0] - UserInfo[0][1] + UserInfo[0][2] - UserInfo[0][3] + TranzitInfo[0][0])
                else:
                    return (UserInfo[0][0] - UserInfo[0][1] + UserInfo[0][2] - UserInfo[0][3])
    else:
        return 0

